# ดึงทุก tab ของทั้ง 2 workbook ออกเป็น JSON + พิมพ์โครงสร้าง
import openpyxl, json

for sf in ['sheet1', 'sheet2']:
    wb = openpyxl.load_workbook(f'scrape/{sf}.xlsx', data_only=True)
    print(f"\n===== {sf}.xlsx — {len(wb.sheetnames)} tabs =====")
    out = {}
    for ws in wb.worksheets:
        rows = []
        for r in ws.iter_rows(values_only=True):
            cells = ['' if c is None else str(c).strip() for c in r]
            if any(c != '' for c in cells):
                # ตัด trailing empty cells
                while cells and cells[-1] == '':
                    cells.pop()
                rows.append(cells)
        out[ws.title] = rows
        print(f"  [{ws.title}] {len(rows)} rows")
    json.dump(out, open(f'scrape/{sf}.json', 'w', encoding='utf-8'), ensure_ascii=False, indent=1)
print("\nwrote scrape/sheet1.json, scrape/sheet2.json")
